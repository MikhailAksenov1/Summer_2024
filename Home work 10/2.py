import openpyxl
from openpyxl import Workbook
#wb = Workbook()
#wb.save('file12.xlsx')
a = openpyxl.load_workbook(filename='file12.xlsx')
b1 = a.active
zp = []
s = 0
for i in range(1,b1.max_row):
    if b1.cell(i,1).value is not None:
        zp.append(b1.cell(i,2).value)
        s += b1.cell(i,2).value
    else:
        break
print(sorted(zp, key=lambda f: f[0][1]))
a.create_sheet('n')
b2 = a['n']
for i in range(1,len(zp)+1):
    k = i-1
    b2.cell(i,1).value = zp[k][0]
    b2.cell(i, 2).value = zp[k][1]
b2.cell(len(zp)+1, 1).value = 'Итого:'
b2.cell(len(zp)+1, 2).value = s
a.save('file12.xlsx')
