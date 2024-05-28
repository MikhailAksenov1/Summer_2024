import openpyxl
#from openpyxl import Workbook
#wb = Workbook()
#wb.save('file13.xlsx')
a = openpyxl.load_workbook(filename='file13.xlsx')
b1 = a.active
zp = []
for i in range(1,5):
    zp.append(b1.cell(i,2).value)
a.create_sheet('результат')
b2 = a['результат']
mediana = (zp[len(zp)//2-1]+zp[len(zp)//2])//2
g = [['Максимальное значение',max(zp)], ['Минимальное значение',min(zp)],['Сумма',sum(zp)],['Среднее арифметическое',sum(zp)//len(zp)],['Медиана',mediana]]
for i in range(1,len(g)+1):
    k = i-1
    b2.cell(i,1).value = g[k][0]
    b2.cell(i, 2).value = g[k][1]
a.save('file13.xlsx')
